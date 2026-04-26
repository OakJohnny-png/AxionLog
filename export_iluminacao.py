import pandas as pd
from sqlalchemy import create_engine, text
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Configurações do Banco de Dados
DB_USER = os.getenv("POSTGRES_USER", "admin")
DB_PASSWORD = os.getenv("POSTGRES_PASSWORD", "sua_senha_segura")
DB_HOST = os.getenv("POSTGRES_HOST", "localhost")
DB_PORT = os.getenv("POSTGRES_PORT", "5432")
DB_NAME = os.getenv("POSTGRES_DB", "logistica_db")

DATABASE_URL = f"postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

def export_to_custom_excel(output_path: str, config: dict = None):
    """
    Exporta os pontos de iluminação pendentes para um Excel formatado para impressão.
    config: dicionário com cores e tamanhos de fonte (baseado na UI do usuário)
    """
    if config is None:
        config = {
            "rota_bg": "FFC000", # Âmbar padrão
            "rota_font_size": 14,
            "bairro_bg": "E2EFDA",
            "bairro_font_size": 12,
            "problema_font_size": 11
        }

    engine = create_engine(DATABASE_URL)
    query = "SELECT rota, bairro, descricao_problema FROM IluminacaoPublicaPontos WHERE status = 'pendente' ORDER BY rota, bairro"
    
    with engine.connect() as conn:
        df = pd.read_sql(query, conn)

    if df.empty:
        print("Não há dados pendentes para exportar.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Rotas de Reparo"

    # Configuração de página para Paisagem
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1

    current_row = 1
    last_rota = None
    last_bairro = None

    # Estilos
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for _, row in df.iterrows():
        # Cabeçalho de Rota
        if row['rota'] != last_rota:
            current_row += 1
            cell = ws.cell(row=current_row, column=1, value=f"--- {row['rota']} ---")
            cell.font = Font(bold=True, size=config['rota_font_size'])
            cell.fill = PatternFill(start_color=config['rota_bg'], end_color=config['rota_bg'], fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            last_rota = row['rota']
            last_bairro = None # Reset bairro ao mudar rota
            current_row += 1

        # Cabeçalho de Bairro
        if row['bairro'] != last_bairro:
            cell = ws.cell(row=current_row, column=1, value=f"Bairro: {row['bairro']}")
            cell.font = Font(bold=True, size=config['bairro_font_size'])
            cell.fill = PatternFill(start_color=config['bairro_bg'], end_color=config['bairro_bg'], fill_type="solid")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            last_bairro = row['bairro']
            current_row += 1

        # Linha do Problema
        ws.cell(row=current_row, column=1, value="[ ]") # Checkbox para o técnico
        ws.cell(row=current_row, column=2, value=row['descricao_problema']).font = Font(size=config['problema_font_size'])
        
        # Bordas para facilitar leitura
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).border = thin_border
        
        current_row += 1

    # Ajuste de largura de colunas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 80

    wb.save(output_path)
    print(f"Planilha de rotas gerada com sucesso: {output_path}")

if __name__ == "__main__":
    # Exemplo de uso
    # export_to_custom_excel("rotas_para_impressao.xlsx")
    print("Para usar, chame export_to_custom_excel('caminho/de/saida.xlsx')")
